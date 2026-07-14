"""
Cihaz Yönetimi sekmesi
"""

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QComboBox, QMessageBox, QLineEdit, QSpinBox,
    QDialog, QTextEdit, QProgressBar, QStyledItemDelegate, QFileDialog
)
from PySide6.QtGui import QColor, QFont
from PySide6.QtCore import Qt, QCoreApplication, QTimer
from datetime import datetime
import time
from models import Device, Employee, get_session
from core.hanvon_client import HanvonClient
from services.employee_sync_service import mark_pending
from services.device_push_worker import DevicePushWorker
from ui.dialogs.device_transfer_dialog import DeviceTransferDialog
import logging

logger = logging.getLogger("HanvonAgent.DeviceMgmt")

# SYNC sütunu renkleri
SYNC_OK_COLOR = QColor(0, 200, 0)       # yeşil — senkron
SYNC_PENDING_COLOR = QColor(255, 255, 0)  # sarı — düzenlendi


class SyncStatusDelegate(QStyledItemDelegate):
    """Sync sütunu için renkli boyama — stylesheet override sorununu aşar."""

    def paint(self, painter, option, index):
        value = index.data(Qt.DisplayRole) or ''

        if value == 'Senkron':
            painter.fillRect(option.rect, SYNC_OK_COLOR)
        elif value == 'Düzenlendi':
            painter.fillRect(option.rect, SYNC_PENDING_COLOR)
        else:
            super().paint(painter, option, index)
            return

        painter.setPen(QColor(0, 0, 0))
        painter.drawText(option.rect, Qt.AlignCenter, value)

# Satır alternasyon renkleri
ROW_WHITE = QColor(255, 255, 255)
ROW_LIGHT_GRAY = QColor(245, 245, 245)

# Sütun index'leri
COL_NUM = 0
COL_ID = 1
COL_NAME = 2
COL_CARD = 3
COL_TYPE = 4
COL_SYNC = 5
COL_ACTIONS = 6


class DeviceMgmtTab(QWidget):
    """Cihaz Yönetimi sekmesi."""

    def __init__(self):
        super().__init__()
        self.session = get_session()
        self.current_employees = []
        self.current_device_id = None
        self._active_workers = []  # Çalışan worker referanslarını tut (GC önle)

        # Sorting state
        self.sort_column = None  # Hangi sütun sıralanıyor
        self.sort_ascending = True  # True = A->Z, False = Z->A

        self._init_ui()
        self._refresh_device_combo()

    def _init_ui(self):
        """Device Mgmt layout."""
        layout = QVBoxLayout(self)

        # Cihaz seçim
        device_group = QGroupBox("Cihaz Seç")
        device_layout = QHBoxLayout()

        device_layout.addWidget(QLabel("Cihaz:"))
        self.device_combo = QComboBox()
        self.device_combo.currentIndexChanged.connect(self._on_device_selected)
        device_layout.addWidget(self.device_combo)

        device_layout.addStretch()

        self.fetch_all_employees_btn = QPushButton("Personelleri Getir")
        self.fetch_all_employees_btn.clicked.connect(self._fetch_all_employees)
        self.fetch_all_employees_btn.setEnabled(False)
        self.fetch_all_employees_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; }")
        device_layout.addWidget(self.fetch_all_employees_btn)

        self.transfer_btn = QPushButton("🔄 Cihaz Transferi")
        self.transfer_btn.setStyleSheet("QPushButton { background-color: #2196F3; color: white; font-weight: bold; }")
        self.transfer_btn.clicked.connect(self._open_transfer_dialog)
        device_layout.addWidget(self.transfer_btn)

        device_group.setLayout(device_layout)
        layout.addWidget(device_group)

        # Personel Tablosu
        self.emp_group = QGroupBox("Personeller Listesi")
        emp_layout = QVBoxLayout()

        # ESKI: Genel filtre (opsiyonel — kaldırabilirsiniz)
        # filter_layout = QHBoxLayout()
        # filter_layout.addWidget(QLabel("Ara:"))
        # self.employee_search = QLineEdit()
        # emp_layout.addLayout(filter_layout)

        # YENİ: Column-based filtering
        self.filter_id = QLineEdit()
        self.filter_id.setPlaceholderText("ID")
        self.filter_id.setMaximumWidth(80)
        self.filter_id.textChanged.connect(self._filter_employees)

        self.filter_name = QLineEdit()
        self.filter_name.setPlaceholderText("İsim")
        self.filter_name.textChanged.connect(self._filter_employees)

        self.filter_card = QLineEdit()
        self.filter_card.setPlaceholderText("Kart No")
        self.filter_card.setMaximumWidth(120)
        self.filter_card.textChanged.connect(self._filter_employees)

        self.filter_type = QLineEdit()
        self.filter_type.setPlaceholderText("Tür")
        self.filter_type.setMaximumWidth(100)
        self.filter_type.textChanged.connect(self._filter_employees)

        self.filter_sync = QComboBox()
        self.filter_sync.addItem("Tümü", None)
        self.filter_sync.addItem("Düzenlendi", "yeni")  # "yeni" = pending
        self.filter_sync.addItem("Senkron", "ok")
        self.filter_sync.setMaximumWidth(100)
        self.filter_sync.currentIndexChanged.connect(self._filter_employees)

        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("🔍 Filtrele:"))
        filter_layout.addWidget(QLabel("ID:"))
        filter_layout.addWidget(self.filter_id)
        filter_layout.addWidget(QLabel("İsim:"))
        filter_layout.addWidget(self.filter_name)
        filter_layout.addWidget(QLabel("Kart:"))
        filter_layout.addWidget(self.filter_card)
        filter_layout.addWidget(QLabel("Tür:"))
        filter_layout.addWidget(self.filter_type)
        filter_layout.addWidget(QLabel("Sync:"))
        filter_layout.addWidget(self.filter_sync)
        filter_layout.addStretch()

        self.bulk_send_btn = QPushButton("📤 Toplu Gönder")
        self.bulk_send_btn.setStyleSheet("QPushButton { background-color: #FF9800; color: white; font-weight: bold; }")
        self.bulk_send_btn.clicked.connect(self._bulk_send_employees)
        self.bulk_send_btn.setEnabled(False)
        filter_layout.addWidget(self.bulk_send_btn)

        self.export_btn = QPushButton("📊 Export")
        self.export_btn.setStyleSheet("QPushButton { background-color: #4CAF50; color: white; font-weight: bold; }")
        self.export_btn.clicked.connect(self._export_employees_to_excel)
        self.export_btn.setEnabled(False)
        filter_layout.addWidget(self.export_btn)

        emp_layout.addLayout(filter_layout)

        self.employee_table = QTableWidget()
        self.employee_table.setColumnCount(7)
        self.employee_table.setHorizontalHeaderLabels([
            "#", "ID", "İsim Bilgisi", "Kart No", "Tür", "sync", "İşlemler"
        ])
        self.employee_table.setColumnWidth(COL_NUM, 40)
        self.employee_table.setColumnWidth(COL_ID, 60)
        self.employee_table.setColumnWidth(COL_NAME, 220)
        self.employee_table.setColumnWidth(COL_CARD, 120)
        self.employee_table.setColumnWidth(COL_TYPE, 100)
        self.employee_table.setColumnWidth(COL_SYNC, 80)
        self.employee_table.setColumnWidth(COL_ACTIONS, 130)

        # Tablo stili (records_tab paleti ile uyumlu)
        self.employee_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                gridline-color: #E0E0E0;
                selection-background-color: #e8e8e8;
                selection-color: #000000;
            }
            QHeaderView::section {
                background-color: #2a2a2a;
                color: white;
                padding: 6px;
                border: none;
                font-weight: bold;
            }
            QTableWidget::item {
                padding: 5px;
                border: none;
            }
            QTableWidget::item:hover {
                background-color: #e8e8e8;
            }
        """)

        # Alternating row colors
        self.employee_table.setAlternatingRowColors(True)

        # Sync sütunu delegate — stylesheet'i bypass eder, renk garantili
        self.employee_table.setItemDelegateForColumn(COL_SYNC, SyncStatusDelegate(self))

        # Header click for sorting
        self.employee_table.horizontalHeader().sectionClicked.connect(self._on_header_click)

        # İnline edit: hücre değiştiğinde mark_pending tetiklenir
        self.employee_table.cellChanged.connect(self._on_cell_changed)

        emp_layout.addWidget(self.employee_table)
        emp_layout.setStretch(1, 1)  # Tabloyu uzat

        self.emp_group.setLayout(emp_layout)
        layout.addWidget(self.emp_group)

    def _refresh_device_combo(self):
        """Cihaz combo'sunu yenile."""
        devices = self.session.query(Device).all()
        self.device_combo.blockSignals(True)
        self.device_combo.clear()
        self.device_combo.addItem("Cihaz Seçiniz", None)
        for device in devices:
            self.device_combo.addItem(f"{device.name} ({device.ip})", device.id)
        self.device_combo.blockSignals(False)

    def _on_device_selected(self):
        """Cihaz seçildi."""
        device_id = self.device_combo.currentData()
        enabled = device_id is not None

        self.fetch_all_employees_btn.setEnabled(enabled)

        if device_id:
            self._load_employees(device_id)
            # Başlığı güncelle - personel sayısı göster
            employee_count = self.session.query(Employee).filter_by(device_id=device_id).count()
            self.emp_group.setTitle(f"Personeller ({employee_count})")
        else:
            self.emp_group.setTitle("Personeller")

    def _load_employees(self, device_id: int):
        """Cihazın personellerini yükle (cache tutarak)."""
        self.current_device_id = device_id
        self.current_employees = self.session.query(Employee).filter_by(device_id=device_id).all()
        # Filtreleri sıfırla
        self.filter_id.clear()
        self.filter_name.clear()
        self.filter_card.clear()
        self.filter_type.clear()
        self.filter_sync.setCurrentIndex(0)  # "Tümü"
        self._filter_employees()

    def _filter_employees(self):
        """Personel listesini tüm filtre kriterleri ile filtrele ve tabloyu yeniden çiz.

        Tüm setItem/setCellWidget çağrıları blockSignals(True) altında yapılır;
        aksi halde programatik doldurma yanlışlıkla cellChanged tetikler.
        """
        # Filtre değerlerini al
        filter_id = self.filter_id.text().lower().strip()
        filter_name = self.filter_name.text().lower().strip()
        filter_card = self.filter_card.text().lower().strip()
        filter_type = self.filter_type.text().lower().strip()
        filter_sync = self.filter_sync.currentData()  # None veya "yeni" veya "ok"

        # Tüm kriterlere göre filtrele (AND logic)
        filtered = []
        for emp in self.current_employees:
            # ID kontrolü
            if filter_id and filter_id not in str(emp.employee_device_id).lower():
                continue

            # İsim kontrolü
            if filter_name:
                emp_name = (emp.display_name or emp.name or '').lower()
                if filter_name not in emp_name:
                    continue

            # Kart No kontrolü
            if filter_card and filter_card not in (emp.card_num or '').lower():
                continue

            # Tür kontrolü
            if filter_type and filter_type not in (emp.check_type or '').lower():
                continue

            # Sync kontrolü
            emp_sync_status = "yeni" if emp.sync_status == "yeni" else "ok"
            if filter_sync is not None and emp_sync_status != filter_sync:
                continue

            filtered.append(emp)

        # Sıralama uygula
        if self.sort_column is not None:
            filtered = self._sort_employees(filtered, self.sort_column, self.sort_ascending)

        self.employee_table.blockSignals(True)
        try:
            self.employee_table.setRowCount(len(filtered))

            for row, emp in enumerate(filtered):
                row_color = ROW_WHITE if row % 2 == 0 else ROW_LIGHT_GRAY
                is_pending = (emp.sync_status == "yeni")

                # Sıra numarası (salt-okunur)
                row_num_item = QTableWidgetItem(str(row + 1))
                row_num_item.setTextAlignment(Qt.AlignCenter)
                row_num_item.setFlags(row_num_item.flags() & ~Qt.ItemIsEditable)
                self.employee_table.setItem(row, COL_NUM, row_num_item)

                # ID (salt-okunur)
                id_item = QTableWidgetItem(str(emp.employee_device_id))
                id_item.setTextAlignment(Qt.AlignCenter)
                id_item.setFlags(id_item.flags() & ~Qt.ItemIsEditable)
                self.employee_table.setItem(row, COL_ID, id_item)

                # İsim (DÜZENLENEBİLİR) — pending varsa display_name gösterilir
                name_item = QTableWidgetItem(emp.display_name or "—")
                name_item.setFlags(name_item.flags() | Qt.ItemIsEditable)
                # Hangi employee'ye ait olduğunu hücrede sakla
                name_item.setData(Qt.UserRole, emp)
                self.employee_table.setItem(row, COL_NAME, name_item)

                # Kart No (salt-okunur)
                card_item = QTableWidgetItem(emp.card_num or "—")
                card_item.setFlags(card_item.flags() & ~Qt.ItemIsEditable)
                self.employee_table.setItem(row, COL_CARD, card_item)

                # Tür (salt-okunur)
                type_item = QTableWidgetItem(emp.check_type or "—")
                type_item.setTextAlignment(Qt.AlignCenter)
                type_item.setFlags(type_item.flags() & ~Qt.ItemIsEditable)
                self.employee_table.setItem(row, COL_TYPE, type_item)

                # SYNC (renkli, salt-okunur)
                sync_text = "Düzenlendi" if is_pending else "Senkron"
                sync_item = QTableWidgetItem(sync_text)
                sync_item.setTextAlignment(Qt.AlignCenter)
                sync_item.setFlags(sync_item.flags() & ~Qt.ItemIsEditable)
                sync_item.setBackground(SYNC_PENDING_COLOR if is_pending else SYNC_OK_COLOR)
                self.employee_table.setItem(row, COL_SYNC, sync_item)

                # Satır arka plan rengi (SYNC hariç — onun kendi rengi var)
                for col in (COL_NUM, COL_ID, COL_NAME, COL_CARD, COL_TYPE):
                    item = self.employee_table.item(row, col)
                    if item:
                        item.setBackground(row_color)

                # İşlem butonları
                self._build_action_widget(row, emp, row_color, is_pending)
        finally:
            self.employee_table.blockSignals(False)

        # Title'ı filtered count ile güncelle
        filtered_count = len(filtered)
        self.emp_group.setTitle(f"Personeller ({filtered_count})")

        # Bulk send butonunu enable/disable et
        has_pending = any(emp.sync_status == "yeni" for emp in self.current_employees)
        self.bulk_send_btn.setEnabled(has_pending)
        self.export_btn.setEnabled(bool(self.current_employees))

    def _build_action_widget(self, row, emp, row_color, is_pending):
        """Bir satırın işlem butonlarını (✎ düzenle, 📤 gönder, ✕ sil) oluştur."""
        action_layout = QHBoxLayout()
        action_layout.setContentsMargins(0, 0, 0, 0)
        action_layout.setSpacing(4)

        edit_btn = QPushButton("✎")
        edit_btn.setMaximumWidth(32)
        edit_btn.setToolTip("İsmi düzenle")
        edit_btn.clicked.connect(lambda checked, r=row: self._start_edit_name(r))
        action_layout.addWidget(edit_btn)

        # 📤 Gönder — yalnızca bekleyen değişikliği olan satırlarda
        if is_pending:
            send_btn = QPushButton("📤")
            send_btn.setMaximumWidth(32)
            send_btn.setToolTip("Değişikliği cihaza gönder")
            send_btn.setStyleSheet("QPushButton { color: #f57c00; font-weight: bold; }")
            send_btn.clicked.connect(lambda checked, e=emp: self._send_employee_to_device(e))
            action_layout.addWidget(send_btn)

        delete_btn = QPushButton("✕")
        delete_btn.setMaximumWidth(32)
        delete_btn.setToolTip("Sil")
        delete_btn.setStyleSheet("QPushButton { color: red; }")
        delete_btn.clicked.connect(lambda checked, e=emp: self._delete_employee(e))
        action_layout.addWidget(delete_btn)

        action_layout.addStretch()

        action_widget = QWidget()
        action_widget.setLayout(action_layout)
        action_widget.setStyleSheet(f"background-color: {row_color.name()};")
        self.employee_table.setCellWidget(row, COL_ACTIONS, action_widget)

    def _start_edit_name(self, row):
        """✎ butonu — İsim hücresini düzenleme moduna al."""
        item = self.employee_table.item(row, COL_NAME)
        if item is not None:
            self.employee_table.editItem(item)

    def _on_cell_changed(self, row, col):
        """İnline edit tamamlandı — İsim sütunu ise mark_pending çağır."""
        if col != COL_NAME:
            return

        item = self.employee_table.item(row, COL_NAME)
        if item is None:
            return

        emp = item.data(Qt.UserRole)
        if emp is None:
            return

        new_name = item.text().strip()
        old_name = emp.name or "(boş)"

        try:
            mark_pending(self.session, emp, new_name)
            # Audit log
            self._write_audit_log_edit(emp.employee_device_id, old_name, new_name, success=True)
            logger.info(f"[EDIT] Personel düzenlendi: ID={emp.employee_device_id} '{old_name}' → '{new_name}'")
        except ValueError as e:
            QMessageBox.warning(self, "Geçersiz İsim", str(e))
            self._write_audit_log_edit(emp.employee_device_id, old_name, new_name, success=False, error=str(e))
        except Exception as e:
            logger.error("İnline edit kaydedilemedi: %s", e, exc_info=True)
            QMessageBox.critical(self, "Hata", f"İsim kaydedilemedi: {e}")
            self._write_audit_log_edit(emp.employee_device_id, old_name, new_name, success=False, error=str(e))

        # SYNC sütununu ve butonları yeniden çiz
        self._filter_employees()

    def _send_employee_to_device(self, employee):
        """📤 — Bekleyen isim değişikliğini cihaza gönder (QThread ile)."""
        device_id = employee.device_id or self.current_device_id
        if device_id is None:
            QMessageBox.warning(self, "Cihaz Yok", "Hedef cihaz belirlenemedi.")
            return

        worker = DevicePushWorker(employee.id, device_id)
        self._active_workers.append(worker)
        worker.finished.connect(
            lambda success, msg, e=employee, w=worker: self._handle_worker_finished(
                success, msg, e, w
            )
        )
        worker.start()

    def _handle_worker_finished(self, success, msg, employee, worker):
        """Worker bitti — UI'yı güncelle, worker referansını temizle."""
        try:
            self.on_push_finished(success, msg, employee)
        finally:
            if worker in self._active_workers:
                self._active_workers.remove(worker)

    def on_push_finished(self, success, msg, employee):
        """Cihaza gönderme sonucu — başarılıysa yenile, değilse hata göster."""
        if success:
            # Worker farklı bir session'da değişiklik yapmış olabilir;
            # identity map'i flush ederek stale cache okumayı önle.
            self.session.expire_all()
            device_id = employee.device_id or self.current_device_id
            if device_id is not None:
                self._load_employees(device_id)
        else:
            QMessageBox.critical(
                self, "Gönderim Başarısız",
                f"Personel cihaza gönderilemedi:\n\n{msg}"
            )

    def _fetch_all_employees(self):
        """Cihazdan TÜM personelleri getir (progress göstergesi ile)."""
        device_id = self.device_combo.currentData()
        if not device_id:
            return

        device = self.session.query(Device).filter_by(id=device_id).first()

        try:
            logger.info(f"[BAŞLAT] Tüm personelleri getir — Cihaz: {device.ip}")

            client = HanvonClient(device.ip, port=device.port, comm_key=device.comm_key)
            client.connect()
            logger.info(f"[OK] Cihaza bağlandı: {device.ip}")

            # GetEmployeeID ile tüm ID'leri al
            logger.info(f"[ADIM 1] GetEmployeeID() çağrılıyor...")
            ids = client.get_employee_id()
            total = len(ids)
            logger.info(f"[OK] {total} personel ID'si alındı")

            if total == 0:
                QMessageBox.warning(self, "Bilgi", "Cihazda personel yok")
                client.disconnect()
                return

            # Dialog oluştur
            result_dialog = QDialog(self)
            result_dialog.setWindowTitle(f"Personel İndiriliyor — {total} Kişi")
            result_dialog.setGeometry(100, 100, 600, 450)
            result_dialog.setFixedSize(600, 450)  # Resize yapılamaz

            layout = QVBoxLayout(result_dialog)

            # Başlık bilgisi
            info_label = QLabel(
                f"📍 Cihaz: {device.name} ({device.ip})\n"
                f"👥 Toplam: {total} personel\n"
                f"⏱️ Bu işlem 2-5 dakika sürebilir..."
            )
            info_label.setStyleSheet("font-weight: bold; color: #1976D2; margin: 10px;")
            layout.addWidget(info_label)

            # Progress bar
            progress_bar = QProgressBar()
            progress_bar.setMaximum(total)
            progress_bar.setValue(0)
            layout.addWidget(progress_bar)

            # Stats label
            stats_label = QLabel("⏳ Hazırlanıyor...")
            stats_label.setStyleSheet("margin: 5px 10px; font-size: 11px;")
            layout.addWidget(stats_label)

            # Sonuç text
            result_text = QTextEdit()
            result_text.setReadOnly(True)
            result_text.setFont(QFont("Consolas", 9))
            result_text.setText(f"🔄 BAŞLADI — {total} Personel Alınıyor\n")
            layout.addWidget(result_text)

            close_btn = QPushButton("Kapat")
            close_btn.setEnabled(False)
            close_btn.clicked.connect(result_dialog.accept)
            layout.addWidget(close_btn)

            result_dialog.show()
            QCoreApplication.processEvents()

            # Veri çekme başla
            logger.info(f"[ADIM 2] Dialog açılıyor...")
            employees_to_add = []
            updated_count = 0
            failed = 0
            failed_ids = []  # Başarısız ID'ler
            chunk_size = 20
            start_time = time.time()
            processed_count = 0

            logger.info(f"[ADIM 3] Veri çekmeye başlanıyor (chunk_size={chunk_size})...")
            for chunk_idx in range(0, total, chunk_size):
                chunk_ids = ids[chunk_idx:chunk_idx + chunk_size]
                chunk_num = chunk_idx // chunk_size + 1
                total_chunks = (total + chunk_size - 1) // chunk_size

                elapsed = time.time() - start_time
                processed_count = chunk_idx

                # Progress bar güncelle
                progress_bar.setValue(processed_count)

                # Stats güncelle (tek satırda)
                if processed_count > 0:
                    speed = processed_count / elapsed
                    remaining = (total - processed_count) / speed if speed > 0 else 0
                    stats_label.setText(
                        f"İşleneni: {processed_count}/{total} | Hız: {speed:.1f}/sn | Tahmini: {remaining:.0f}s"
                    )
                else:
                    stats_label.setText(f"Başlanıyor...")

                QCoreApplication.processEvents()

                successful_in_chunk = 0
                for emp_id in chunk_ids:
                    try:
                        logger.debug(f"GetEmployee({emp_id}) çağrılıyor...")
                        emp_data = client.get_employee(emp_id)
                        if emp_data and emp_data.get('result') == 'success':
                            # Mevcut kontrol et
                            existing = self.session.query(Employee).filter_by(
                                device_id=device_id,
                                employee_device_id=int(emp_id)
                            ).first()

                            name = emp_data.get('name', '').strip()
                            card = emp_data.get('card_num', '').strip()
                            # Tüm *_data alanlarını birleştir (face_data, finger_data vb.)
                            biometric_templates = []
                            for k, v in emp_data.items():
                                if k.endswith('_data') and isinstance(v, list):
                                    biometric_templates.extend(v)

                            if existing:
                                existing.name = name
                                existing.card_num = card
                                existing.check_type = emp_data.get('check_type', 'face')
                                existing.authority = emp_data.get('authority', '')
                                existing.calid = emp_data.get('calid', '')
                                existing.opendoor_type = emp_data.get('opendoor_type', 'face')
                                existing.last_synced = datetime.utcnow()
                                if biometric_templates:
                                    existing.face_data = biometric_templates
                                updated_count += 1
                            else:
                                emp = Employee(
                                    employee_device_id=int(emp_id),
                                    name=name,
                                    card_num=card,
                                    check_type=emp_data.get('check_type', 'face'),
                                    authority=emp_data.get('authority', ''),
                                    calid=emp_data.get('calid', ''),
                                    opendoor_type=emp_data.get('opendoor_type', 'face'),
                                    device_id=device_id,
                                )
                                if biometric_templates:
                                    emp.face_data = biometric_templates
                                employees_to_add.append(emp)
                            successful_in_chunk += 1
                    except Exception as e:
                        failed += 1
                        failed_ids.append(int(emp_id))
                        logger.debug(f"GetEmployee({emp_id}) başarısız: {str(e)}")
                        continue

                # Blok sonucu göster
                result_text.append(
                    f"📦 [Blok {chunk_num}/{total_chunks}] ID {int(chunk_ids[0]):03d}-{int(chunk_ids[-1]):03d} → {successful_in_chunk}/{len(chunk_ids)} OK"
                )
                QCoreApplication.processEvents()

            # DB'ye kaydet
            logger.info(f"[ADIM 4] Veritabanına kaydediliyor...")
            self.session.add_all(employees_to_add)
            self.session.commit()
            logger.info(f"[OK] {len(employees_to_add)} yeni, {updated_count} güncellendi")

            client.disconnect()
            logger.info(f"[OK] Bağlantı kapatıldı")

            # Final sonuç
            elapsed_total = time.time() - start_time
            progress_bar.setValue(total)

            result_text.append("\n" + "=" * 70)
            result_text.append("✅ TAMAMLANDI!")
            result_text.append(f"✨ Yeni: {len(employees_to_add)} | 🔄 Güncellenen: {updated_count} | ❌ Başarısız: {failed} | ⏱️ {elapsed_total:.1f}s")

            # Başarısız ID'ler varsa göster
            if failed_ids:
                failed_str = ', '.join(map(str, sorted(failed_ids)))
                result_text.append(f"\n❌ Başarısız ID'ler: {failed_str}")
                logger.warning(f"[BAŞARISIZ] {len(failed_ids)} personel yüklenemedi: {failed_ids}")

            # DB'de var, cihazda yok kontrolü
            device_id_set = {str(i) for i in ids}
            all_db_emps = self.session.query(Employee).filter_by(device_id=device_id).all()
            orphaned = [e for e in all_db_emps if str(e.employee_device_id) not in device_id_set]
            result_text.append("\n" + "─" * 70)
            if orphaned:
                result_text.append(f"⚠️  DB'DE VAR — CİHAZDA YOK: {len(orphaned)} personel")
                result_text.append(f"   (DB: {len(all_db_emps)} kayıt  |  Cihaz: {len(ids)} ID)")
                result_text.append("")
                for e in sorted(orphaned, key=lambda x: x.employee_device_id):
                    result_text.append(f"   ID {e.employee_device_id:>5}  {e.name or '—'}")
                logger.warning(
                    "[FARK] DB'de var cihazda yok: %s",
                    [e.employee_device_id for e in orphaned],
                )
            else:
                result_text.append(
                    f"✅ DB ve cihaz eşleşiyor  (DB: {len(all_db_emps)}  |  Cihaz: {len(ids)})"
                )

            stats_label.setText(f"✅ Tamamlandı — {elapsed_total:.1f}s sürdü")
            stats_label.setStyleSheet("margin: 5px 10px; font-size: 11px; color: green; font-weight: bold;")

            logger.info(f"[TAMAMLANDI] {total} personel işlendi ({elapsed_total:.1f}s)")
            logger.info(f"  Yeni: {len(employees_to_add)}, Güncellenen: {updated_count}, Başarısız: {failed}")

            close_btn.setEnabled(True)
            result_dialog.exec()
            self._load_employees(device_id)

        except Exception as e:
            logger.error(f"[HATA] İşlem başarısız: {str(e)}", exc_info=True)
            error_msg = f"""İşlem başarısız oldu:

{str(e)}

Lütfen:
1. Cihaz IP'sinin doğru olduğunu kontrol edin
2. Cihazın açık ve bağlı olduğunu kontrol edin
3. Network bağlantısını kontrol edin
4. Tekrar deneyin

Debug: Konsol çıktısını kontrol edin"""
            QMessageBox.critical(self, "HATA", error_msg)
        finally:
            try:
                if 'client' in locals() and client:
                    client.disconnect()
                    logger.info("[OK] Bağlantı kapatıldı")
            except:
                pass

    def _delete_employee(self, employee):
        """Personeli sil (çift onay + cihaz + DB + audit log)."""
        name = employee.display_name or str(employee.employee_device_id)
        emp_id = employee.employee_device_id

        # ONAY 1: DB'den silme
        reply1 = QMessageBox.question(
            self, "Personeli Sil - Adım 1/2",
            f"'{name}' (ID={emp_id}) personeli VERİTABANINDAN silinecek.\n\nEmin misiniz?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply1 != QMessageBox.Yes:
            return

        # ONAY 2: Cihazdan da silme mi?
        reply2 = QMessageBox.question(
            self, "Personeli Sil - Adım 2/2",
            f"Personel CİHAZDAN DA kaldirilacak.\n\n{name} (ID={emp_id}) cihazdan kaldirılacak. Emin misiniz?",
            QMessageBox.Yes | QMessageBox.No,
        )

        # Silme işlemi başla
        db_deleted = False
        device_deleted = False
        device_error = None

        try:
            # 1. DB'den sil
            self.session.delete(employee)
            self.session.commit()
            db_deleted = True
            logger.info(f"[DB] Personel silindi: ID={emp_id}, Name={name}")

            # 2. reply2 == Yes ise cihazdan da sil
            if reply2 == QMessageBox.Yes:
                try:
                    device = self.session.query(Device).filter_by(id=self.current_device_id).first()
                    if device:
                        client = HanvonClient(device.ip, port=device.port, comm_key=device.comm_key)
                        client.connect()

                        result = client.delete_employee(str(emp_id))
                        client.disconnect()

                        if result:
                            device_deleted = True
                            logger.info(f"[CIHAZ] Personel silindi: ID={emp_id}, Device={device.ip}")
                        else:
                            device_error = "Cihaz yanıt vermedi"
                            logger.warning(f"[CIHAZ] Silme başarısız: ID={emp_id}, Device={device.ip}")
                except Exception as e:
                    device_error = str(e)
                    logger.error(f"[CIHAZ] Hata: {device_error}")

            # 3. Audit log yaz
            self._write_audit_log(emp_id, name, db_deleted, device_deleted, device_error)

            # 4. UI güncelle
            device_id = self.device_combo.currentData()
            self._load_employees(device_id)

            # 5. Final mesaj
            if device_deleted:
                QMessageBox.information(
                    self, "Başarılı",
                    f"Personel DB'den silindi, cihazdan kaldırıldı.\n\n{name} (ID={emp_id})"
                )
            else:
                if device_error:
                    QMessageBox.warning(
                        self, "Kısmi Başarı",
                        f"Personel DB'den silindi, cihazda korundu.\n\nHata: {device_error}"
                    )
                else:
                    QMessageBox.information(
                        self, "Başarılı",
                        f"Personel DB'den silindi.\n(Cihazda kaldı)"
                    )

        except Exception as e:
            logger.error(f"[DELETE] Hata: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "Hata", f"Silme başarısız:\n{str(e)}")

    def _write_audit_log(self, emp_id, emp_name, db_deleted, device_deleted, device_error):
        """Silme işlemini audit log'a yaz."""
        try:
            from datetime import datetime
            from core import app_paths

            log_file = app_paths.logs_dir() / "audit.log"
            log_file.parent.mkdir(parents=True, exist_ok=True)

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            device = self.session.query(Device).filter_by(id=self.current_device_id).first()
            device_ip = device.ip if device else "Unknown"

            status = []
            if db_deleted:
                status.append("DB:OK")
            if device_deleted:
                status.append("CIHAZ:OK")
            elif device_error:
                status.append(f"CIHAZ:FAILED({device_error})")
            else:
                status.append("CIHAZ:SKIP")

            log_line = f"[{timestamp}] DELETE: ID={emp_id} Name={emp_name} Device={device_ip} {' '.join(status)}\n"

            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(log_line)

            logger.debug(f"[AUDIT] Logged: {log_line.strip()}")

        except Exception as e:
            logger.error(f"[AUDIT] Log yazma hatası: {str(e)}")

    def _write_audit_log_edit(self, emp_id, old_name, new_name, success=True, error=None):
        """Personel edit işlemini audit log'a yaz."""
        try:
            from datetime import datetime
            from core import app_paths

            log_file = app_paths.logs_dir() / "audit.log"
            log_file.parent.mkdir(parents=True, exist_ok=True)

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            device = self.session.query(Device).filter_by(id=self.current_device_id).first()
            device_ip = device.ip if device else "Unknown"

            if success:
                status = "OK"
                log_line = f"[{timestamp}] EDIT: ID={emp_id} '{old_name}' → '{new_name}' Device={device_ip} {status}\n"
            else:
                status = f"FAILED({error})"
                log_line = f"[{timestamp}] EDIT: ID={emp_id} '{old_name}' → '{new_name}' Device={device_ip} {status}\n"

            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(log_line)

            logger.debug(f"[AUDIT] Logged: {log_line.strip()}")

        except Exception as e:
            logger.error(f"[AUDIT] Log yazma hatası: {str(e)}")

    def _sort_employees(self, employees, col, ascending):
        """Personel listesini verilen sütuna göre sırala."""
        if col == COL_ID:
            return sorted(employees, key=lambda e: int(e.employee_device_id or 0), reverse=not ascending)
        elif col == COL_NAME:
            return sorted(employees, key=lambda e: (e.display_name or e.name or '').lower(), reverse=not ascending)
        elif col == COL_CARD:
            return sorted(employees, key=lambda e: (e.card_num or '').lower(), reverse=not ascending)
        elif col == COL_TYPE:
            return sorted(employees, key=lambda e: (e.check_type or '').lower(), reverse=not ascending)
        elif col == COL_SYNC:
            # Yeni (pending) sütunu — "yeni" önce, "ok" sonra
            return sorted(employees, key=lambda e: e.sync_status == "ok", reverse=not ascending)
        else:
            return employees

    def _on_header_click(self, col):
        """Header'a tıklandı — sıralama yap."""
        # Sıralama yapılamayan sütunlar
        if col in (COL_NUM, COL_ACTIONS):
            return

        # Aynı sütuna tekrar tıklandıysa → ters sırala
        if self.sort_column == col:
            self.sort_ascending = not self.sort_ascending
        else:
            self.sort_column = col
            self.sort_ascending = True

        # Tabloyu yeniden çiz (sıralanmış)
        self._filter_employees()

        # Header'ı güncelle (ok işareti göster)
        self._update_header_indicator()

    def _update_header_indicator(self):
        """Header'da sıralama yönünü göster (▲▼)."""
        header = self.employee_table.horizontalHeader()
        headers = [
            "#", "ID", "İsim Bilgisi", "Kart No", "Tür", "sync", "İşlemler"
        ]

        for col, title in enumerate(headers):
            if col == self.sort_column:
                arrow = "▲" if self.sort_ascending else "▼"
                header.model().setHeaderData(col, Qt.Horizontal, f"{title} {arrow}")
            else:
                header.model().setHeaderData(col, Qt.Horizontal, title)

    def _open_transfer_dialog(self):
        """Cihaz transferi dialog'unu aç."""
        dialog = DeviceTransferDialog(self)
        dialog.exec()

    def _bulk_send_employees(self):
        """Düzenlenen personelleri tek tek cihaza gönder (hata toleranslı)."""
        # Düzenlenen personelleri bul
        pending_employees = [
            emp for emp in self.current_employees
            if emp.sync_status == "yeni" and emp.pending_name
        ]

        if not pending_employees:
            QMessageBox.information(self, "Bilgi", "Gönderilecek düzenlenme yok.")
            return

        device_id = self.device_combo.currentData()
        if not device_id:
            QMessageBox.warning(self, "Hata", "Cihaz seçiniz.")
            return

        device = self.session.query(Device).filter_by(id=device_id).first()
        if not device:
            QMessageBox.critical(self, "Hata", "Cihaz bulunamadı.")
            return

        # Onay diyalogu
        names_list = "\n".join([f"  ID {emp.employee_device_id}: {emp.name} → {emp.pending_name}" for emp in pending_employees])
        reply = QMessageBox.question(
            self, "Toplu Gönder (Tek Tek)",
            f"{len(pending_employees)} personel gönderilecek:\n\n{names_list}\n\nEmin misiniz?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        # Progress diyalogu
        progress_dialog = QDialog(self)
        progress_dialog.setWindowTitle("Personeller Gönderiliyor...")
        progress_dialog.setGeometry(100, 100, 500, 300)
        progress_dialog.setFixedSize(500, 300)

        layout = QVBoxLayout(progress_dialog)

        result_text = QTextEdit()
        result_text.setReadOnly(True)
        result_text.setFont(QFont("Consolas", 9))
        layout.addWidget(result_text)

        progress_dialog.show()
        QCoreApplication.processEvents()

        # Tek tek gönder
        successful = 0
        failed_list = []
        client = None

        try:
            client = HanvonClient(device.ip, port=device.port, comm_key=device.comm_key)
            client.connect()
            logger.info(f"[TOPLU] Cihaza bağlandı: {device.ip}")

            for idx, emp in enumerate(pending_employees, 1):
                emp_info = f"{emp.pending_name} (ID {emp.employee_device_id})"
                result_text.append(f"{idx}/{len(pending_employees)}: {emp_info} → Gönderiliyor...")
                QCoreApplication.processEvents()

                try:
                    # Tek bir personel için SetNameTable
                    success = client.set_name_table({
                        str(emp.employee_device_id): emp.pending_name
                    })

                    if success:
                        # DB'yi güncelle
                        emp.name = emp.pending_name
                        emp.pending_name = None
                        emp.sync_status = "ok"
                        self.session.commit()

                        result_text.append(f"   ✅ Başarılı")
                        successful += 1
                        logger.info(f"[TOPLU] Gönderilen: ID {emp.employee_device_id} '{emp.name}'")
                    else:
                        result_text.append(f"   ❌ Cihaz reddetti")
                        failed_list.append(emp)
                        logger.warning(f"[TOPLU] Reddedilen: ID {emp.employee_device_id}")

                except Exception as e:
                    result_text.append(f"   ❌ Hata: {str(e)[:50]}")
                    failed_list.append(emp)
                    logger.debug(f"[TOPLU] Hata (ID {emp.employee_device_id}): {str(e)}")
                    continue

            client.disconnect()
            logger.info(f"[TOPLU] Bağlantı kapatıldı")

        except Exception as e:
            result_text.append(f"\n❌ Bağlantı hatası: {str(e)}")
            logger.error(f"[TOPLU] Bağlantı hatası: {str(e)}", exc_info=True)
            failed_list = pending_employees  # Hepsi başarısız
        finally:
            if client:
                try:
                    client.disconnect()
                except:
                    pass

        # Özet
        result_text.append("\n" + "=" * 50)
        result_text.append(f"✅ Başarılı: {successful}/{len(pending_employees)}")
        if failed_list:
            result_text.append(f"❌ Başarısız: {len(failed_list)} (tekrar göndermek için butona basın)")
            failed_ids = [f"ID {e.employee_device_id}" for e in failed_list]
            result_text.append(f"   {', '.join(failed_ids)}")

        # Audit log
        self._write_audit_log_bulk_send(pending_employees, successful, len(failed_list), device_id)

        # UI güncelle
        self._load_employees(device_id)

        # Kapat butonu
        close_btn = QPushButton("Kapat")
        close_btn.clicked.connect(progress_dialog.accept)
        layout.addWidget(close_btn)

        progress_dialog.exec()

    def _export_employees_to_excel(self):
        """Seçili cihazdaki personelleri XLS olarak dışa aktar."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.critical(self, "Hata", "openpyxl kütüphanesi bulunamadı.\npip install openpyxl")
            return

        if not self.current_employees:
            QMessageBox.information(self, "Bilgi", "Dışa aktarılacak personel yok.")
            return

        device_id = self.device_combo.currentData()
        device = self.session.query(Device).filter_by(id=device_id).first()
        device_label = device.ip if device else f"device_{device_id}"

        default_name = f"personeller_{device_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel Olarak Kaydet", default_name,
            "Excel Dosyası (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Personeller"

        # Başlık satırı
        headers = ["#", "Cihaz ID", "İsim", "Bekleyen İsim", "Kart No", "Tür", "Yetki", "Sync", "Son Senkron", "Oluşturulma"]
        header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Veri satırları
        pending_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        for row_idx, emp in enumerate(self.current_employees, 2):
            row_data = [
                row_idx - 1,
                emp.employee_device_id,
                emp.name or "",
                emp.pending_name or "",
                emp.card_num or "",
                emp.check_type or "",
                emp.authority or "",
                emp.sync_status or "",
                emp.last_synced.strftime("%Y-%m-%d %H:%M:%S") if emp.last_synced else "",
                emp.created_at.strftime("%Y-%m-%d %H:%M:%S") if emp.created_at else "",
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if emp.sync_status == "yeni":
                    cell.fill = pending_fill

        # Sütun genişliklerini otomatik ayarla
        col_widths = [5, 10, 30, 30, 16, 12, 10, 8, 20, 20]
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        try:
            wb.save(path)
            QMessageBox.information(
                self, "Başarılı",
                f"{len(self.current_employees)} personel dışa aktarıldı:\n{path}"
            )
            logger.info(f"[EXPORT] {len(self.current_employees)} personel → {path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Dosya kaydedilemedi:\n{str(e)}")
            logger.error(f"[EXPORT] Kayıt hatası: {e}")

    def _write_audit_log_bulk_send(self, employees, successful_count, failed_count, device_id):
        """Toplu gönderme işlemini audit log'a yaz."""
        try:
            from datetime import datetime
            from core import app_paths

            log_file = app_paths.logs_dir() / "audit.log"
            log_file.parent.mkdir(parents=True, exist_ok=True)

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            device = self.session.query(Device).filter_by(id=device_id).first()
            device_ip = device.ip if device else "Unknown"

            emp_ids = ', '.join(str(e.employee_device_id) for e in employees)
            log_line = f"[{timestamp}] BULK_SEND: {successful_count} OK / {failed_count} FAILED (ID: {emp_ids}) Device={device_ip}\n"

            with open(log_file, 'a', encoding='utf-8') as f:
                f.write(log_line)

            logger.info(f"[AUDIT] Logged: {log_line.strip()}")

        except Exception as e:
            logger.error(f"[AUDIT] Log yazma hatası: {str(e)}")

